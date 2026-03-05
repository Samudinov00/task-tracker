"""
Views для проектов, задач, комментариев и уведомлений.

Управление доступом:
  manager  — полный CRUD проектов и задач; видит все задачи своих проектов
  executor — редактирует только свои задачи (статус + комментарии)
  client   — только чтение задач своих проектов + комментарии
"""
import csv
import io
import json
from datetime import timedelta

from django.db.models import Count, Q, Sum
from django.contrib import messages
from django.core.cache import cache
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DeleteView, ListView, UpdateView

from accounts.models import CustomUser
from .forms import AttachmentForm, CommentForm, ProjectForm, TaskForm, TaskStatusForm, TimeLogForm
from .models import (
    Comment, Notification, Project, Task, TaskAttachment,
    TaskChangeLog, TaskStatusLog, TimeLog,
)
from .tasks import send_notifications


# ─────────────────────────── Хелперы ──────────────────────────────────────────

class ManagerRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_manager()

    def handle_no_permission(self):
        if not self.request.user.is_authenticated:
            return redirect('accounts:login')
        messages.error(self.request, 'Доступ запрещён. Требуются права менеджера.')
        return redirect('projects:project_list')


def _check_project_access(user, project):
    """Выбрасывает PermissionDenied, если пользователь не имеет доступа к проекту."""
    if user.is_manager():
        if project.manager != user:
            raise PermissionDenied
    elif user.is_client():
        if not user.client_projects.filter(pk=project.pk).exists():
            raise PermissionDenied
    elif user.is_executor():
        # Исполнитель видит проект если он в списке команды ИЛИ назначен на задачу
        in_team = user.executor_projects.filter(pk=project.pk).exists()
        has_task = project.tasks.filter(assignee=user).exists()
        if not (in_team or has_task):
            raise PermissionDenied


def _notify(users, task, ntype, message):
    """Ставит уведомления в очередь Celery."""
    if not hasattr(users, '__iter__') or isinstance(users, str):
        users = [users]
    user_ids = [u.pk for u in users if u]
    if user_ids:
        send_notifications.delay(user_ids, task.pk, ntype, message)



# ─────────────────────────── Проекты ──────────────────────────────────────────

class ProjectListView(LoginRequiredMixin, ListView):
    model = Project
    template_name = 'projects/project_list.html'
    context_object_name = 'projects'

    def get_queryset(self):
        user = self.request.user
        if user.is_manager():
            return Project.objects.filter(manager=user)
        if user.is_client():
            return user.client_projects.all()
        if user.is_executor():
            # Проекты из команды + проекты с назначенными задачами
            return Project.objects.filter(
                Q(executors=user) | Q(tasks__assignee=user)
            ).distinct()
        return Project.objects.none()


class ProjectCreateView(ManagerRequiredMixin, CreateView):
    model = Project
    form_class = ProjectForm
    template_name = 'projects/project_form.html'

    def form_valid(self, form):
        form.instance.manager = self.request.user
        messages.success(self.request, f'Проект «{form.instance.name}» создан.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('projects:kanban', kwargs={'project_uuid': self.object.uuid})

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Новый проект'
        return ctx


class ProjectUpdateView(ManagerRequiredMixin, UpdateView):
    model = Project
    form_class = ProjectForm
    template_name = 'projects/project_form.html'

    def get_object(self, queryset=None):
        obj = get_object_or_404(Project, uuid=self.kwargs['uuid'])
        if obj.manager != self.request.user:
            raise PermissionDenied
        return obj

    def get_success_url(self):
        return reverse('projects:project_detail', kwargs={'uuid': self.object.uuid})

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Редактировать проект'
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Проект обновлён.')
        return super().form_valid(form)


class ProjectDeleteView(ManagerRequiredMixin, DeleteView):
    model = Project
    template_name = 'projects/project_confirm_delete.html'
    success_url = reverse_lazy('projects:project_list')

    def get_object(self, queryset=None):
        obj = get_object_or_404(Project, uuid=self.kwargs['uuid'])
        if obj.manager != self.request.user:
            raise PermissionDenied
        return obj

    def form_valid(self, form):
        messages.success(self.request, 'Проект удалён.')
        return super().form_valid(form)


@login_required
def project_detail(request, uuid):
    project = get_object_or_404(Project, uuid=uuid)
    _check_project_access(request.user, project)

    tasks = project.tasks.select_related('assignee')
    if request.user.is_executor():
        tasks = tasks.filter(assignee=request.user)
    elif request.user.is_client():
        tasks = tasks.filter(clients=request.user)
    ctx = {
        'project': project,
        'tasks': tasks,
        'not_started_count': tasks.filter(status='not_started').count(),
        'in_work_count': tasks.filter(status='development').count(),
        'test_count': tasks.filter(status__in=['test_nsk', 'test_district']).count(),
        'production_count': tasks.filter(status='production').count(),
    }
    return render(request, 'projects/project_detail.html', ctx)


# ─────────────────────────── Канбан-доска ─────────────────────────────────────

@login_required
def kanban(request, project_uuid):
    project = get_object_or_404(Project, uuid=project_uuid)
    _check_project_access(request.user, project)

    qs = project.tasks.select_related('assignee', 'project').prefetch_related('comments')
    if request.user.is_executor():
        qs = qs.filter(assignee=request.user)
    elif request.user.is_client():
        qs = qs.filter(clients=request.user)

    assignee_filter = None
    priority_filter = None
    deadline_filter = None
    executors = []

    if request.user.is_manager():
        executors = list(
            CustomUser.objects.filter(assigned_tasks__project=project).distinct().order_by('first_name', 'username')
        )
        assignee_id = request.GET.get('assignee')
        if assignee_id == 'none':
            qs = qs.filter(assignee__isnull=True)
            assignee_filter = 'none'
        elif assignee_id:
            qs = qs.filter(assignee_id=assignee_id)
            assignee_filter = assignee_id

    # Priority filter (available to all)
    priority = request.GET.get('priority')
    if priority in [c[0] for c in Task.PRIORITY_CHOICES]:
        qs = qs.filter(priority=priority)
        priority_filter = priority

    # Deadline filter
    deadline = request.GET.get('deadline')
    now = timezone.localdate()
    if deadline == 'overdue':
        qs = qs.filter(deadline__lt=now).exclude(status=Task.STATUS_PRODUCTION)
        deadline_filter = 'overdue'
    elif deadline == 'soon':
        qs = qs.filter(deadline__gte=now, deadline__lte=now + timedelta(days=7)).exclude(status=Task.STATUS_PRODUCTION)
        deadline_filter = 'soon'

    kanban_columns = [
        {
            'key': 'not_started', 'label': 'Не начата',
            'color': 'secondary', 'icon': 'bi-circle',
            'tasks': qs.filter(status='not_started').order_by('order'),
        },
        {
            'key': 'development', 'label': 'Разработка',
            'color': 'primary', 'icon': 'bi-code-slash',
            'tasks': qs.filter(status='development').order_by('order'),
        },
        {
            'key': 'test_nsk', 'label': 'Тест НСК',
            'color': 'warning', 'icon': 'bi-bug',
            'tasks': qs.filter(status='test_nsk').order_by('order'),
        },
        {
            'key': 'test_district', 'label': 'Тест район',
            'color': 'warning', 'icon': 'bi-geo-alt-fill',
            'tasks': qs.filter(status='test_district').order_by('order'),
        },
        {
            'key': 'production', 'label': 'Промышленная эксплуатация',
            'color': 'success', 'icon': 'bi-check-circle-fill',
            'tasks': qs.filter(status='production').order_by('order'),
        },
    ]
    return render(request, 'projects/kanban.html', {
        'project': project,
        'kanban_columns': kanban_columns,
        'executors': executors,
        'assignee_filter': assignee_filter,
        'priority_filter': priority_filter,
        'deadline_filter': deadline_filter,
        'priority_choices': Task.PRIORITY_CHOICES,
        'status_choices': Task.STATUS_CHOICES,
    })


# ─────────────── AJAX: перемещение задачи в Канбане ───────────────────────────

@login_required
@require_POST
def task_move(request, task_uuid):
    task = get_object_or_404(Task, uuid=task_uuid)
    user = request.user

    # Клиент не может двигать карточки
    if user.is_client():
        return JsonResponse({'error': 'Нет доступа'}, status=403)
    # Исполнитель — только свои задачи
    if user.is_executor() and task.assignee != user:
        return JsonResponse({'error': 'Нет доступа'}, status=403)
    # Менеджер — только свои проекты
    if user.is_manager() and task.project.manager != user:
        return JsonResponse({'error': 'Нет доступа'}, status=403)

    data = json.loads(request.body)
    new_status  = data.get('status')
    column_ids  = data.get('column_ids', [])   # упорядоченный список UUID задач в колонке

    valid_statuses = [s[0] for s in Task.STATUS_CHOICES]
    if new_status not in valid_statuses:
        return JsonResponse({'error': 'Неверный статус'}, status=400)

    old_status   = task.status
    task.status  = new_status
    task.save(update_fields=['status', 'updated_at'])

    # Обновить порядок всей колонки
    for idx, tuuid in enumerate(column_ids):
        Task.objects.filter(uuid=tuuid, project=task.project).update(order=idx)

    # Уведомления и лог при смене статуса
    if old_status != new_status:
        TaskStatusLog.objects.create(task=task, changed_by=user, old_status=old_status, new_status=new_status)
        label = dict(Task.STATUS_CHOICES).get(new_status)
        recipients = set()
        if task.assignee and task.assignee != user:
            recipients.add(task.assignee)
        if task.project.manager != user:
            recipients.add(task.project.manager)
        _notify(
            recipients, task, Notification.TYPE_TASK_STATUS,
            f'Статус задачи «{task.title}» изменён на «{label}»',
        )

    return JsonResponse({'success': True})


# ─────────────── AJAX: состояние канбан-доски (polling) ───────────────────────

@login_required
def kanban_state_api(request, project_uuid):
    project = get_object_or_404(Project, uuid=project_uuid)
    _check_project_access(request.user, project)

    qs = project.tasks.select_related('assignee')
    if request.user.is_executor():
        qs = qs.filter(assignee=request.user)
    elif request.user.is_client():
        qs = qs.filter(clients=request.user)

    tasks_data = []
    for task in qs:
        tasks_data.append({
            'uuid': str(task.uuid),
            'title': task.title,
            'status': task.status,
            'priority': task.priority,
            'assignee_name': task.assignee.get_display_name() if task.assignee else '',
            'deadline': task.deadline.isoformat() if task.deadline else None,
            'is_overdue': task.is_overdue(),
            'comment_count': task.comments.count(),
        })

    return JsonResponse({'tasks': tasks_data, 'updated_at': timezone.now().isoformat()})


# ─────────────────────────── Задачи ───────────────────────────────────────────

@login_required
def task_detail(request, uuid):
    task = get_object_or_404(
        Task.objects.select_related('assignee', 'project', 'created_by')
                    .prefetch_related('comments__author', 'time_logs__user', 'change_logs__changed_by'),
        uuid=uuid,
    )
    user = request.user
    _check_project_access(user, task.project)

    # Исполнитель видит только свои задачи
    if user.is_executor() and task.assignee != user:
        raise PermissionDenied
    # Клиент видит только задачи где он указан в поле clients
    if user.is_client() and not task.clients.filter(pk=user.pk).exists():
        raise PermissionDenied

    comment_form    = CommentForm()
    attachment_form = AttachmentForm()
    timelog_form    = TimeLogForm()

    can_edit     = user.is_manager() or (user.is_executor() and task.assignee == user)
    can_log_time = can_edit

    if request.method == 'POST':
        if 'comment_submit' in request.POST:
            comment_form = CommentForm(request.POST)
            if comment_form.is_valid():
                comment = comment_form.save(commit=False)
                comment.task   = task
                comment.author = user
                comment.save()

                # Уведомления
                recipients = set()
                if task.assignee and task.assignee != user:
                    recipients.add(task.assignee)
                if task.project.manager != user:
                    recipients.add(task.project.manager)
                _notify(
                    recipients, task, Notification.TYPE_COMMENT,
                    f'{user.get_display_name()} прокомментировал задачу «{task.title}»',
                )

                messages.success(request, 'Комментарий добавлен.')
                return redirect('projects:task_detail', uuid=task.uuid)

        elif 'attachment_submit' in request.POST:
            if not user.is_manager():
                raise PermissionDenied
            attachment_form = AttachmentForm(request.POST, request.FILES)
            if attachment_form.is_valid():
                TaskAttachment.objects.create(
                    task=task,
                    file=attachment_form.cleaned_data['file'],
                    uploaded_by=user,
                )
                messages.success(request, 'Файл загружен.')
                return redirect('projects:task_detail', uuid=task.uuid)

        elif 'timelog_submit' in request.POST:
            if not can_log_time:
                raise PermissionDenied
            timelog_form = TimeLogForm(request.POST)
            if timelog_form.is_valid():
                log = timelog_form.save(commit=False)
                log.task = task
                log.user = user
                log.save()
                messages.success(request, 'Время добавлено.')
                return redirect('projects:task_detail', uuid=task.uuid)

    # Объединить историю изменений: TaskStatusLog + TaskChangeLog
    status_logs = list(task.status_logs.select_related('changed_by').all())
    change_logs = list(task.change_logs.select_related('changed_by').all())

    # Конвертируем статус-логи в формат совместимый с change_logs для отображения
    history = []
    for log in status_logs:
        old_label = dict(Task.STATUS_CHOICES).get(log.old_status, log.old_status)
        new_label = dict(Task.STATUS_CHOICES).get(log.new_status, log.new_status)
        history.append({
            'type': 'status',
            'changed_by': log.changed_by,
            'field_name': 'Статус',
            'old_value': old_label,
            'new_value': new_label,
            'changed_at': log.changed_at,
        })
    for log in change_logs:
        history.append({
            'type': 'field',
            'changed_by': log.changed_by,
            'field_name': log.field_name,
            'old_value': log.old_value,
            'new_value': log.new_value,
            'changed_at': log.changed_at,
        })
    history.sort(key=lambda x: x['changed_at'], reverse=True)

    # Итоги по времени
    total_minutes = task.get_total_logged_minutes()
    total_hours   = total_minutes // 60
    remaining_min = total_minutes % 60

    return render(request, 'projects/task_detail.html', {
        'task': task,
        'comment_form': comment_form,
        'attachment_form': attachment_form,
        'timelog_form': timelog_form,
        'can_edit': can_edit,
        'can_log_time': can_log_time,
        'history': history,
        'total_hours': total_hours,
        'remaining_min': remaining_min,
        'total_minutes': total_minutes,
    })


@login_required
def task_create(request, project_uuid):
    if not request.user.is_manager():
        raise PermissionDenied

    project = get_object_or_404(Project, uuid=project_uuid, manager=request.user)

    if request.method == 'POST':
        form = TaskForm(project=project, data=request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.project    = project
            task.created_by = request.user
            task.save()
            form.save_m2m()

            if task.assignee:
                _notify(
                    task.assignee, task, Notification.TYPE_TASK_ASSIGNED,
                    f'Вам назначена задача «{task.title}» в проекте «{project.name}»',
                )

            messages.success(request, f'Задача «{task.title}» создана.')
            return redirect('projects:kanban', project_uuid=project.uuid)
    else:
        form = TaskForm(project=project)

    return render(request, 'projects/task_form.html', {
        'form': form,
        'project': project,
        'title': 'Новая задача',
    })


@login_required
def task_edit(request, uuid):
    task = get_object_or_404(Task, uuid=uuid)
    user = request.user
    project = task.project

    if user.is_client():
        raise PermissionDenied
    if user.is_executor() and task.assignee != user:
        raise PermissionDenied
    if user.is_manager() and project.manager != user:
        raise PermissionDenied

    # Исполнитель меняет только статус/под-этап
    FormClass = TaskStatusForm if user.is_executor() else TaskForm
    form_kwargs = {} if user.is_executor() else {'project': project}

    if request.method == 'POST':
        form = FormClass(**form_kwargs, data=request.POST, instance=task)
        if form.is_valid():
            old_status = task.status
            task._changed_by = user
            saved_task = form.save()

            if user.is_manager():
                # Уведомить исполнителя при смене назначения
                if saved_task.assignee and 'assignee' in form.changed_data:
                    _notify(
                        saved_task.assignee, saved_task, Notification.TYPE_TASK_ASSIGNED,
                        f'Вам назначена задача «{saved_task.title}»',
                    )

            if old_status != saved_task.status:
                TaskStatusLog.objects.create(task=saved_task, changed_by=user, old_status=old_status, new_status=saved_task.status)
                label = dict(Task.STATUS_CHOICES).get(saved_task.status)
                recipients = set()
                if saved_task.assignee and saved_task.assignee != user:
                    recipients.add(saved_task.assignee)
                if project.manager != user:
                    recipients.add(project.manager)
                _notify(
                    recipients, saved_task, Notification.TYPE_TASK_STATUS,
                    f'Статус задачи «{saved_task.title}» изменён на «{label}»',
                )

            messages.success(request, 'Задача обновлена.')
            return redirect('projects:task_detail', uuid=saved_task.uuid)
    else:
        form = FormClass(**form_kwargs, instance=task)

    return render(request, 'projects/task_form.html', {
        'form': form,
        'task': task,
        'project': project,
        'title': 'Редактировать задачу',
    })


@login_required
def task_delete(request, uuid):
    task = get_object_or_404(Task, uuid=uuid)
    if not request.user.is_manager() or task.project.manager != request.user:
        raise PermissionDenied

    project_uuid = task.project.uuid
    if request.method == 'POST':
        task.delete()
        messages.success(request, 'Задача удалена.')
        return redirect('projects:kanban', project_uuid=project_uuid)

    return render(request, 'projects/task_confirm_delete.html', {'task': task})


# ─────────────────────────── Логи статусов ────────────────────────────────────

@login_required
def status_logs(request):
    if not request.user.is_manager():
        raise PermissionDenied

    logs = TaskStatusLog.objects.select_related(
        'task__project', 'changed_by'
    ).filter(task__project__manager=request.user)

    # Фильтр по проекту
    project_id = request.GET.get('project')
    if project_id:
        logs = logs.filter(task__project_id=project_id)

    # Фильтр по исполнителю
    executor_id = request.GET.get('executor')
    if executor_id:
        logs = logs.filter(changed_by_id=executor_id)

    projects = Project.objects.filter(manager=request.user)
    executors = CustomUser.objects.filter(
        status_changes__task__project__manager=request.user
    ).distinct()

    return render(request, 'projects/status_logs.html', {
        'logs': logs[:200],
        'projects': projects,
        'executors': executors,
        'selected_project': project_id,
        'selected_executor': executor_id,
    })


# ─────────────────────────── Уведомления ──────────────────────────────────────

@login_required
def notifications_list(request):
    notifications = request.user.notifications.select_related('task__project').all()
    # Помечаем все как прочитанные при открытии страницы
    request.user.notifications.filter(is_read=False).update(is_read=True)
    return render(request, 'projects/notifications.html', {
        'notifications': notifications,
    })


@login_required
def notifications_count_api(request):
    count = request.user.notifications.filter(is_read=False).count()
    return JsonResponse({'count': count})


@login_required
def notifications_recent_api(request):
    items = request.user.notifications.filter(is_read=False).select_related('task')[:6]
    data = []
    for n in items:
        data.append({
            'id': n.pk,
            'message': n.message,
            'created_at': n.created_at.strftime('%d.%m %H:%M'),
            'task_id': n.task.pk if n.task else None,
            'task_url': reverse('projects:task_detail', kwargs={'uuid': n.task.uuid}) if n.task else None,
        })
    total = request.user.notifications.filter(is_read=False).count()
    return JsonResponse({'notifications': data, 'count': total})


@login_required
@require_POST
def notifications_mark_all_read(request):
    request.user.notifications.filter(is_read=False).update(is_read=True)
    return JsonResponse({'success': True})


# ─────────────────────────── Вложения ─────────────────────────────────────────

@login_required
@require_POST
def attachment_delete(request, pk):
    attachment = get_object_or_404(TaskAttachment, pk=pk)
    user = request.user
    task = attachment.task

    # Только менеджер проекта или загрузивший файл могут удалять
    is_manager = user.is_manager() and task.project.manager == user
    is_uploader = attachment.uploaded_by == user
    if not (is_manager or is_uploader):
        raise PermissionDenied

    # Удаляем файл с диска
    if attachment.file and attachment.file.storage.exists(attachment.file.name):
        attachment.file.delete(save=False)
    attachment.delete()

    messages.success(request, 'Вложение удалено.')
    return redirect('projects:task_detail', uuid=task.uuid)


# ─────────────────────────── Трекер времени ───────────────────────────────────

@login_required
@require_POST
def log_time(request, task_uuid):
    task = get_object_or_404(Task, uuid=task_uuid)
    user = request.user

    can_log = user.is_manager() or (user.is_executor() and task.assignee == user)
    if not can_log:
        return JsonResponse({'error': 'Нет доступа'}, status=403)

    form = TimeLogForm(request.POST)
    if form.is_valid():
        log = form.save(commit=False)
        log.task = task
        log.user = user
        log.save()
        messages.success(request, f'Добавлено {log.minutes} мин.')
    else:
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error)

    return redirect('projects:task_detail', uuid=task.uuid)


# ─────────────────────────── Bulk операции ────────────────────────────────────

@login_required
@require_POST
def bulk_task_update(request):
    if not request.user.is_manager():
        raise PermissionDenied

    task_uuids = request.POST.getlist('task_uuids[]')
    action     = request.POST.get('action')
    value      = request.POST.get('value')

    if not task_uuids or not action or not value:
        messages.error(request, 'Некорректные данные.')
        return redirect(request.META.get('HTTP_REFERER', 'projects:project_list'))

    tasks = Task.objects.filter(
        uuid__in=task_uuids,
        project__manager=request.user,
    ).select_related('assignee', 'project')

    updated = 0
    if action == 'change_status':
        valid_statuses = [s[0] for s in Task.STATUS_CHOICES]
        if value not in valid_statuses:
            messages.error(request, 'Неверный статус.')
            return redirect(request.META.get('HTTP_REFERER', 'projects:project_list'))

        label = dict(Task.STATUS_CHOICES).get(value)
        for task in tasks:
            old_status = task.status
            if old_status != value:
                task.status = value
                task.save(update_fields=['status', 'updated_at'])
                TaskStatusLog.objects.create(
                    task=task, changed_by=request.user,
                    old_status=old_status, new_status=value,
                )
                TaskChangeLog.objects.create(
                    task=task, changed_by=request.user,
                    field_name='Статус',
                    old_value=dict(Task.STATUS_CHOICES).get(old_status, old_status),
                    new_value=label,
                )
                updated += 1

    elif action == 'change_assignee':
        try:
            new_assignee = CustomUser.objects.get(pk=value, role='executor')
        except CustomUser.DoesNotExist:
            messages.error(request, 'Исполнитель не найден.')
            return redirect(request.META.get('HTTP_REFERER', 'projects:project_list'))

        for task in tasks:
            old_assignee_name = task.assignee.get_display_name() if task.assignee else '—'
            task.assignee = new_assignee
            task.save(update_fields=['assignee', 'updated_at'])
            TaskChangeLog.objects.create(
                task=task, changed_by=request.user,
                field_name='Исполнитель',
                old_value=old_assignee_name,
                new_value=new_assignee.get_display_name(),
            )
            _notify(
                new_assignee, task, Notification.TYPE_TASK_ASSIGNED,
                f'Вам назначена задача «{task.title}»',
            )
            updated += 1

    if updated:
        messages.success(request, f'Обновлено задач: {updated}.')
    else:
        messages.info(request, 'Нет изменений.')

    return redirect(request.META.get('HTTP_REFERER', 'projects:project_list'))


# ─────────────────────────── Аналитика ────────────────────────────────────────

@login_required
def analytics(request):
    if not request.user.is_manager():
        raise PermissionDenied

    projects = Project.objects.filter(manager=request.user)
    project_filter = request.GET.get('project')
    selected_project = None
    if project_filter:
        try:
            selected_project = projects.get(uuid=project_filter)
        except Project.DoesNotExist:
            pass

    cache_key = f'analytics:{request.user.pk}:{project_filter or "all"}'
    ctx = cache.get(cache_key)

    if ctx is None:
        tasks_qs = Task.objects.filter(project__manager=request.user).select_related('assignee', 'project')
        if selected_project:
            tasks_qs = tasks_qs.filter(project=selected_project)

        status_labels = [label for _, label in Task.STATUS_CHOICES]
        status_data   = [tasks_qs.filter(status=key).count() for key, _ in Task.STATUS_CHOICES]

        priority_labels = [label for _, label in Task.PRIORITY_CHOICES]
        priority_data   = [tasks_qs.filter(priority=key).count() for key, _ in Task.PRIORITY_CHOICES]

        assignee_stats = (
            tasks_qs
            .exclude(assignee__isnull=True)
            .values('assignee__first_name', 'assignee__last_name', 'assignee__username')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        assignee_labels = []
        assignee_data   = []
        for row in assignee_stats:
            name = f"{row['assignee__first_name']} {row['assignee__last_name']}".strip() or row['assignee__username']
            assignee_labels.append(name)
            assignee_data.append(row['count'])

        today = timezone.localdate()
        dates = [today - timedelta(days=i) for i in range(29, -1, -1)]
        created_labels = [d.strftime('%d.%m') for d in dates]
        created_data = [tasks_qs.filter(created_at__date=d).count() for d in dates]

        total_tasks      = tasks_qs.count()
        total_projects   = projects.count()
        production_count = tasks_qs.filter(status=Task.STATUS_PRODUCTION).count()
        overdue_count    = tasks_qs.filter(
            deadline__lt=today
        ).exclude(status=Task.STATUS_PRODUCTION).count()

        ctx = {
            'total_tasks': total_tasks,
            'total_projects': total_projects,
            'production_count': production_count,
            'overdue_count': overdue_count,
            'status_labels_json':   json.dumps(status_labels,   ensure_ascii=False),
            'status_data_json':     json.dumps(status_data),
            'priority_labels_json': json.dumps(priority_labels, ensure_ascii=False),
            'priority_data_json':   json.dumps(priority_data),
            'assignee_labels_json': json.dumps(assignee_labels, ensure_ascii=False),
            'assignee_data_json':   json.dumps(assignee_data),
            'created_labels_json':  json.dumps(created_labels,  ensure_ascii=False),
            'created_data_json':    json.dumps(created_data),
        }
        cache.set(cache_key, ctx, timeout=300)  # 5 минут

    ctx['projects'] = projects
    ctx['selected_project'] = selected_project
    return render(request, 'projects/analytics.html', ctx)


# ─────────────────────────── Экспорт ──────────────────────────────────────────

def _get_export_tasks(request):
    """Возвращает queryset задач менеджера с оптимизированными JOIN."""
    return (
        Task.objects
        .filter(project__manager=request.user)
        .select_related('project', 'assignee')
        .prefetch_related('clients', 'time_logs')
        .annotate(total_minutes=Sum('time_logs__minutes'))
        .order_by('project__name', 'status', 'title')
    )


@login_required
def export_tasks_csv(request):
    if not request.user.is_manager():
        raise PermissionDenied

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="tasks.csv"'

    writer = csv.writer(response)
    writer.writerow(['Название', 'Проект', 'Статус', 'Приоритет', 'Исполнитель', 'Дедлайн', 'Клиенты', 'Время (мин)'])

    status_map   = dict(Task.STATUS_CHOICES)
    priority_map = dict(Task.PRIORITY_CHOICES)

    for task in _get_export_tasks(request):
        clients_str = ', '.join(c.get_display_name() for c in task.clients.all())
        writer.writerow([
            task.title,
            task.project.name,
            status_map.get(task.status, task.status),
            priority_map.get(task.priority, task.priority),
            task.assignee.get_display_name() if task.assignee else '',
            task.deadline.strftime('%d.%m.%Y') if task.deadline else '',
            clients_str,
            task.total_minutes or 0,
        ])

    return response


@login_required
def export_tasks_excel(request):
    if not request.user.is_manager():
        raise PermissionDenied

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'Библиотека openpyxl не установлена. Установите: pip install openpyxl')
        return redirect('projects:analytics')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Задачи'

    # Заголовки
    headers = ['Название', 'Проект', 'Статус', 'Приоритет', 'Исполнитель', 'Дедлайн', 'Клиенты', 'Время (мин)']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    status_map   = dict(Task.STATUS_CHOICES)
    priority_map = dict(Task.PRIORITY_CHOICES)

    for row_idx, task in enumerate(_get_export_tasks(request), 2):
        clients_str = ', '.join(c.get_display_name() for c in task.clients.all())
        ws.cell(row=row_idx, column=1, value=task.title)
        ws.cell(row=row_idx, column=2, value=task.project.name)
        ws.cell(row=row_idx, column=3, value=status_map.get(task.status, task.status))
        ws.cell(row=row_idx, column=4, value=priority_map.get(task.priority, task.priority))
        ws.cell(row=row_idx, column=5, value=task.assignee.get_display_name() if task.assignee else '')
        ws.cell(row=row_idx, column=6, value=task.deadline.strftime('%d.%m.%Y') if task.deadline else '')
        ws.cell(row=row_idx, column=7, value=clients_str)
        ws.cell(row=row_idx, column=8, value=task.total_minutes or 0)

    # Авто-ширина колонок
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="tasks.xlsx"'
    return response
