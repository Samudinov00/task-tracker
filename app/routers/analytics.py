"""
Роуты аналитики и экспорта (аналог соответствующих views в projects/views.py).
"""
import csv
import io
import json
import uuid as uuid_lib
from datetime import date, timedelta

from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.dependencies import get_db, get_unread_count, require_manager
from app.models.project import Project, Task, PRIORITY_CHOICES
from app.models.user import User
from app.utils import templates

router = APIRouter()


def _get_export_tasks(db: Session, user: User):
    return (
        db.query(Task)
        .join(Project, Task.project_id == Project.id)
        .filter(Project.manager_id == user.id)
        .options(
            joinedload(Task.project),
            joinedload(Task.assignee),
            joinedload(Task.status_obj),
        )
        .order_by(Project.name, Task.title)
        .all()
    )


@router.get("/analytics/", response_class=HTMLResponse, name="analytics")
async def analytics(request: Request, db: Session = Depends(get_db)):
    user = require_manager(request, db)
    projects = db.query(Project).filter(Project.manager_id == user.id).all()

    project_filter = request.query_params.get("project")
    selected_project = None
    if project_filter:
        try:
            selected_project = next((p for p in projects if str(p.uuid) == project_filter), None)
        except Exception:
            pass

    tasks_q = db.query(Task).join(Project).filter(Project.manager_id == user.id)
    if selected_project:
        tasks_q = tasks_q.filter(Task.project_id == selected_project.id)

    all_tasks = tasks_q.options(joinedload(Task.assignee), joinedload(Task.status_obj)).all()

    status_labels = ["В работе", "Завершено"]
    status_data = [
        sum(1 for t in all_tasks if not (t.status_obj and t.status_obj.is_final)),
        sum(1 for t in all_tasks if t.status_obj and t.status_obj.is_final),
    ]

    priority_labels = [label for _, label in PRIORITY_CHOICES]
    priority_data = [sum(1 for t in all_tasks if t.priority == key) for key, _ in PRIORITY_CHOICES]

    # Нагрузка по исполнителям (топ-10)
    assignee_counts: dict = {}
    for t in all_tasks:
        if t.assignee:
            name = t.assignee.get_display_name()
            assignee_counts[name] = assignee_counts.get(name, 0) + 1
    sorted_assignees = sorted(assignee_counts.items(), key=lambda x: -x[1])[:10]
    assignee_labels = [a[0] for a in sorted_assignees]
    assignee_data = [a[1] for a in sorted_assignees]

    today = date.today()
    dates = [today - timedelta(days=i) for i in range(29, -1, -1)]
    created_labels = [d.strftime("%d.%m") for d in dates]
    created_data = [sum(1 for t in all_tasks if t.created_at and t.created_at.date() == d) for d in dates]

    total_tasks = len(all_tasks)
    total_projects = len(projects)
    production_count = sum(1 for t in all_tasks if t.status_obj and t.status_obj.is_final)
    overdue_count = sum(1 for t in all_tasks if t.deadline and t.deadline < today and not (t.status_obj and t.status_obj.is_final))

    unread_count = get_unread_count(db, user.id)
    return templates.TemplateResponse(
        "projects/analytics.html",
        {
            "request": request, "user": user,
            "projects": projects, "selected_project": selected_project,
            "total_tasks": total_tasks, "total_projects": total_projects,
            "production_count": production_count, "overdue_count": overdue_count,
            "status_labels_json": json.dumps(status_labels, ensure_ascii=False),
            "status_data_json": json.dumps(status_data),
            "priority_labels_json": json.dumps(priority_labels, ensure_ascii=False),
            "priority_data_json": json.dumps(priority_data),
            "assignee_labels_json": json.dumps(assignee_labels, ensure_ascii=False),
            "assignee_data_json": json.dumps(assignee_data),
            "created_labels_json": json.dumps(created_labels, ensure_ascii=False),
            "created_data_json": json.dumps(created_data),
            "unread_notifications_count": unread_count,
        },
    )


@router.get("/export/tasks/csv/", name="export_tasks_csv")
async def export_tasks_csv(request: Request, db: Session = Depends(get_db)):
    user = require_manager(request, db)
    tasks = _get_export_tasks(db, user)

    priority_map = dict(PRIORITY_CHOICES)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Название", "Проект", "Статус", "Приоритет", "Исполнитель", "Дедлайн", "Клиенты"])

    for task in tasks:
        clients_str = ", ".join(c.get_display_name() for c in task.clients)
        writer.writerow([
            task.title,
            task.project.name,
            task.get_status_display(),
            priority_map.get(task.priority, task.priority),
            task.assignee.get_display_name() if task.assignee else "",
            task.deadline.strftime("%d.%m.%Y") if task.deadline else "",
            clients_str,
        ])

    output.seek(0)
    # UTF-8 BOM для корректного открытия в Excel
    content = "\ufeff" + output.getvalue()
    return StreamingResponse(
        iter([content.encode("utf-8")]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="tasks.csv"'},
    )


@router.get("/export/tasks/excel/", name="export_tasks_excel")
async def export_tasks_excel(request: Request, db: Session = Depends(get_db)):
    user = require_manager(request, db)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi.responses import RedirectResponse as RR
        return RR(url="/analytics/", status_code=302)

    tasks = _get_export_tasks(db, user)
    priority_map = dict(PRIORITY_CHOICES)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Задачи"

    headers = ["Название", "Проект", "Статус", "Приоритет", "Исполнитель", "Дедлайн", "Клиенты"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, task in enumerate(tasks, 2):
        clients_str = ", ".join(c.get_display_name() for c in task.clients)
        ws.cell(row=row_idx, column=1, value=task.title)
        ws.cell(row=row_idx, column=2, value=task.project.name)
        ws.cell(row=row_idx, column=3, value=task.get_status_display())
        ws.cell(row=row_idx, column=4, value=priority_map.get(task.priority, task.priority))
        ws.cell(row=row_idx, column=5, value=task.assignee.get_display_name() if task.assignee else "")
        ws.cell(row=row_idx, column=6, value=task.deadline.strftime("%d.%m.%Y") if task.deadline else "")
        ws.cell(row=row_idx, column=7, value=clients_str)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tasks.xlsx"'},
    )
